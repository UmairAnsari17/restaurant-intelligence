from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_excel(restaurants, city):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = city

    headers = [
        "Restaurant Name",
        "Phone Number",
        "Address",
        "Cuisine",
        "Google Maps"
    ]

    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    # Header Row
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data Rows
    for row, restaurant in enumerate(restaurants, start=2):

        sheet.cell(row=row, column=1).value = restaurant.get("name", "")
        sheet.cell(row=row, column=2).value = restaurant.get("phone", "")
        sheet.cell(row=row, column=3).value = restaurant.get("address", "")
        sheet.cell(row=row, column=4).value = restaurant.get("cuisine", "")

        link_cell = sheet.cell(row=row, column=5)
        link = restaurant.get("google_maps", "")

        link_cell.value = "Open Map"
        link_cell.hyperlink = link
        link_cell.style = "Hyperlink"

    # Auto-fit columns
    for column_cells in sheet.columns:

        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )

        sheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = length + 5

    # Freeze header
    sheet.freeze_panes = "A2"

    filename = f"{city}_restaurants.xlsx"

    workbook.save(filename)

    return filename